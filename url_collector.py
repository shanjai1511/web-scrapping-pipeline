import os
import sys
import yaml
import importlib.util
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import time
import hashlib
import json
from common_module import CommonModule
base_dir = "C:/Users/shanj/OneDrive/Desktop/web-scrapping-pipeline"

class UrlCollector(CommonModule):
    def __init__(self, base_dir):
        self.base_dir = base_dir
        self.output_dir = ""
        self.collector_dir = ""
        self.project_name = ""
        self.site_name = ""
        self.count = 0

    def write_url_in_txt(self, result_url):
        filepath = os.path.join(self.output_dir, f"{self.site_name}_{self.project_name}.txt")
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'a') as file:
            for item in result_url:
                file.write(f"{item}\n")
        self.print_status("success", filepath, "Successfully written URLs")

    def print_status(self, status, file_name, info):
        status_message = {
            "status": status,
            "file_name": file_name,
            "project": self.project_name,
            "site_name": self.site_name,
            "info": info
        }
        json_message = json.dumps(status_message, indent=4)
        print(json_message)

    def encode(self, array):
        combined_str = ''.join(array)
        unique_id = hashlib.md5(combined_str.encode()).hexdigest()
        return unique_id

    def enter_count_in_sheet(self):
        excel_file = os.path.join(self.base_dir, "url_collector", "url_collector_count_sheet.xlsx")
        sheet_name = "collector_count"
        book = load_workbook(excel_file)
        sheet = book[sheet_name]

        row_num = 2
        while sheet.cell(row=row_num, column=1).value is not None:
            row_num += 1

        array = [time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), self.site_name, self.project_name]
        unique_id = self.encode(array)
        total_url_count = self.count

        sheet.cell(row=row_num, column=1, value=unique_id)
        sheet.cell(row=row_num, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        sheet.cell(row=row_num, column=3, value=self.project_name)
        sheet.cell(row=row_num, column=4, value=self.site_name)
        sheet.cell(row=row_num, column=5, value=total_url_count)

        book.save(excel_file)
        print(f"Data written to {excel_file}")

    def get_final_url(self, url, depth, current_depth_level, max_depth, module_instance):
        current_depth = depth[f"depth{current_depth_level}"]
        method_name = current_depth["method_name"]
        method_to_call = getattr(module_instance, method_name)
        if method_to_call is None:
            self.print_status("error", "", f"Method {method_name} not found in module instance.")
            return

        for i in url:
            try:
                self.print_status("processing", "", f"Processing {i} with method {method_name}")
                result_url = method_to_call(i, depth, current_depth_level)
            except Exception as e:
                self.print_status("error", "", f"Error calling method {method_name} with argument {i}: {e}")
                continue
            if current_depth_level == max_depth:
                self.write_url_in_txt(result_url)
                self.count += len(result_url)
            else:
                self.get_final_url(result_url, depth, current_depth_level + 1, max_depth, module_instance)

    def main(self):
        if len(sys.argv) != 3:
            self.print_status("error", "", "Usage: python url_collector.py <project_name> <site_name>")
            sys.exit(1)

        self.project_name = sys.argv[1]
        self.site_name = sys.argv[2]
        self.output_dir = os.path.join(self.base_dir, f"scrape_output/collector_output/{self.project_name}")
        self.collector_dir = os.path.join(self.base_dir, f"url_collector/{self.project_name}")

        filepath = os.path.join(self.output_dir, f"{self.site_name}_{self.project_name}.txt")
        with open(filepath, 'w') as file:
            file.write('')
        self.print_status("success", filepath, "Cleared existing output file")

        self.main_execution()

    def main_execution(self):
        try:
            self.print_status("info", "", "Starting script execution.")
            yaml_file_path = os.path.join(self.collector_dir, f"{self.site_name}_{self.project_name}.yml")
            self.print_status("info", "", f"Loading configuration file: {yaml_file_path}")
            with open(yaml_file_path, 'r') as file:
                depth = yaml.safe_load(file)

            module_path = os.path.join(self.collector_dir, f"{self.site_name}_{self.project_name}.py")
            self.print_status("info", "", f"Loading module: {module_path}")
            class_name_in_site_script = f"{self.site_name}_{self.project_name}"
            class_name_in_site_script = ''.join([word.capitalize() for word in class_name_in_site_script.split('_')])
            
            try:
                self.print_status("info", "", f"Importing module from {module_path}")
                spec = importlib.util.spec_from_file_location(class_name_in_site_script, module_path)
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                SiteClass = getattr(module, class_name_in_site_script)
                site_instance = SiteClass()  # Create an instance of the dynamically loaded class
            except Exception as e:
                self.print_status("error", "", f"Error importing module from {module_path}: {e}")
                return

            seed_url = depth["depth0"]["seed_url"]
            if not isinstance(seed_url, list):
                seed_url = [seed_url]

            self.get_final_url(seed_url, depth, 0, len(depth) - 1, site_instance)  # Pass the instance
            self.enter_count_in_sheet()

        except Exception as e:
            self.print_status("error", "", f"Unhandled error during execution: {e}")
            raise

if __name__ == "__main__":
    collector = UrlCollector(base_dir)
    collector.main()
