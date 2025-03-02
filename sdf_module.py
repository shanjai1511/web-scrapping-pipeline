import os
import glob
import requests
import hashlib
import json
import sys
import csv
import yaml
import importlib.util
import logging
from openpyxl import load_workbook  # type: ignore
from datetime import datetime
import time
from bs4 import BeautifulSoup
from pathlib import Path
from datetime import date
from lxml import etree #type: ignore
from proxy_config import *
import random
from lxml import html #type: ignore
import json
import math
import pdb

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler("C:/Users/shanj/OneDrive/Desktop/web-scrapping-pipeline/logs/pipeline.log"),
        logging.StreamHandler(sys.stdout)
    ]
)

class CommonModule:
    @staticmethod
    def print_error_message(status, info):
        status_message = {
            "status": status,
            "info": info
        }
        json_message = json.dumps(status_message, indent=4)
        logging.error(json_message)

    @staticmethod
    def print_info_message(status, info=None, url=None):
        status_message = {
            "status": status,
            "info": info
        }
        if url is not None:
            status_message["url"] = url
            
        json_message = json.dumps(status_message, indent=4)
        logging.info(json_message)

    @staticmethod
    def get_page_content_hash(url, proxy=None, extended_header=None):
        if url:
            try:
                CommonModule.print_info_message("info", f"Fetching page content for URL: {url}")
                # Prepare the proxies dictionary if Webshare proxy credentials are provided
                if proxy == "webshare_proxy":
                    host, port, username, password = random.choice(webshare_proxy)
                    proxy_url = f"http://{username}:{password}@{host}:{port}"
                    proxies = {"http": proxy_url, "https": proxy_url}

                    # Make the request with headers and proxies
                    response = requests.get(
                        url,
                        headers=extended_header,
                        verify=False,
                        proxies=proxies
                    )
                else:
                    response = requests.get(
                        url,
                        headers=extended_header,
                        verify=False
                    )

                if response.status_code == 200:
                    result = {
                        "page_doc": response.text,
                        "status_code": response.status_code,
                        "url": url
                    }
                    output_dir = Path("C:/Users/shanj/OneDrive/Desktop/web-scrapping-pipeline/cache/")
                    output_dir.mkdir(parents=True, exist_ok=True)

                    output_file = output_dir / f"{CommonModule.encode(url)}.html"
                    with open(output_file, 'wb') as file:
                        file.write(response.content)
                    CommonModule.print_info_message("success", str(output_file), "Page fetched successfully.")
                    return result
                else:
                    CommonModule.print_error_message("error", f"Failed to fetch page content for URL: {url}")
                    return {
                        "page_doc": response.text,
                        "status_code": response.status_code,
                        "url": url
                    }

            except requests.RequestException as e:
                CommonModule.print_error_message("error", f"Request failed for URL: {url} with error: {e}")
                logging.exception("Request failed")
                return {
                    "page_doc": "",
                    "status_code": None,
                    "url": url
                }
        else:
            CommonModule.print_error_message("error", "Invalid URL")
            return {"page_doc": "", "status_code": None, "url": url}

    @staticmethod
    def get_parsed_tree(page_doc, format="lxml"):
        try:
            CommonModule.print_info_message("info", f"Parsing page document using {format}")
            if format == "lxml":
                parsed_tree = html.fromstring(page_doc["page_doc"])
                CommonModule.print_info_message("success", "Page document parsed successfully using lxml.")
                return parsed_tree
            else:
                if type(page_doc) == dict:
                    page_doc = page_doc["page_doc"]
                soup = BeautifulSoup(page_doc, 'html5lib')
                CommonModule.print_info_message("success", "Page document parsed successfully using beautiful soup.")
                return soup
        except Exception as e:
            CommonModule.print_error_message("error", f"Unexpected error during parsing: {e}")
            logging.exception("Parsing failed")
            return None

    @staticmethod
    def get_value_from_xpath(parsed_tree, xpath_expr, count, attr="none"):
        try:
            CommonModule.print_info_message("info", f"Extracting value from XPath expression: {xpath_expr}")
            elements = parsed_tree.select(xpath_expr)
            text_content = [element.get_text() for element in elements if element]
            if attr != "none":
                text_content = [link[attr] for link in elements if link.has_attr(attr)]
            CommonModule.print_info_message("success", f"Value extracted from XPath expression: {xpath_expr}")
            return text_content if count == "all" else (text_content[0] if text_content else None)
        except Exception as e:
            CommonModule.print_error_message("error", f"XPath extraction failed with error: {e}")
            logging.exception("XPath extraction failed")
            return f"Unexpected error: {e}"

    @staticmethod
    def get_value_from_css_selector(parsed_tree, css_selector, count, attr="none"):
        try:
            CommonModule.print_info_message("info", f"Extracting value from CSS selector: {css_selector}")
            elements = parsed_tree.select(css_selector)
            text_content = [element.get_text() for element in elements if element]
            if attr != "none":
                text_content = [element.get(attr) for element in elements if element.has_attr(attr)]
            CommonModule.print_info_message("success", f"Value extracted from CSS selector: {css_selector}")
            return text_content if count == "all" else (text_content[0] if text_content else None)
        except Exception as e:
            CommonModule.print_error_message("error", f"CSS selector extraction failed with error: {e}")
            logging.exception("CSS selector extraction failed")
            return f"Unexpected error: {e}"

    @staticmethod
    def encode(array):
        CommonModule.print_info_message("info", "Encoding array")
        combined_str = ''.join(array)
        unique_id = hashlib.md5(combined_str.encode()).hexdigest()
        CommonModule.print_info_message("success", "Array encoded successfully")
        return unique_id

class UrlCollector:
    def __init__(self, base_dir, project_name, site_name):
        CommonModule.print_info_message("info", f"Initializing UrlCollector for project: {project_name} and site: {site_name}")
        self.base_dir = base_dir
        self.output_dir = ""
        self.collector_dir = ""
        self.project_name = project_name
        self.site_name = site_name
        self.count = 0

    def write_url_in_txt(self, result_url):
        CommonModule.print_info_message("info", f"Writing URLs to file for project: {self.project_name} and site: {self.site_name}")
        filepath = Path(self.output_dir) / f"{self.site_name}_{self.project_name}.txt"
        filepath.parent.mkdir(parents=True, exist_ok=True)
        with open(filepath, 'a') as file:
            for item in result_url:
                file.write(f"{item}\n")
        CommonModule.print_info_message("success", f"URLs written to file successfully for project: {self.project_name} and site: {self.site_name}")

    def enter_count_in_sheet(self):
        CommonModule.print_info_message("info", f"Entering count in sheet for project: {self.project_name} and site: {self.site_name}")
        if self.count <= 0:
            CommonModule.print_info_message("Failure","No URLs to enter in the sheet")
            return
        excel_file = Path(self.base_dir) / "url_collector" / "url_collector_count_sheet.xlsx"
        sheet_name = "collector_count"
        book = load_workbook(excel_file)
        sheet = book[sheet_name]

        row_num = 2
        while sheet.cell(row=row_num, column=1).value is not None:
            row_num += 1

        array = [time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), self.site_name, self.project_name]
        unique_id = CommonModule.encode(array)
        total_url_count = self.count

        sheet.cell(row=row_num, column=1, value=unique_id)
        sheet.cell(row=row_num, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        sheet.cell(row=row_num, column=3, value=self.project_name)
        sheet.cell(row=row_num, column=4, value=self.site_name)
        sheet.cell(row=row_num, column=5, value=total_url_count)

        book.save(excel_file)
        CommonModule.print_info_message("success", f"Data written to sheet successfully for project: {self.project_name} and site: {self.site_name}")

    def get_final_url(self, url, depth, current_depth_level, max_depth, module_instance):
        CommonModule.print_info_message("info", f"Getting final URL for project: {self.project_name} and site: {self.site_name}")
        current_depth = depth[f"depth{current_depth_level}"]
        method_name = current_depth["method_name"]
        method_to_call = getattr(module_instance, method_name)
        if method_to_call is None:
            return
        
        for i in url:
            try:
                result_url = method_to_call(i, depth, current_depth_level)
            except Exception as e:
                CommonModule.print_error_message("error", f"URL fetching failed with error: {e}")
                logging.exception("URL fetching failed")
                continue
            if current_depth_level == max_depth:
                self.write_url_in_txt(result_url)
                self.count += len(result_url)
            else:
                self.get_final_url(result_url, depth, current_depth_level + 1, max_depth, module_instance)

    def main_execution(self):
        CommonModule.print_info_message("info", f"Starting main execution for project: {self.project_name} and site: {self.site_name}")
        try:
            yaml_file_path = Path(self.collector_dir) / f"{self.site_name}_{self.project_name}.yml"
            CommonModule.print_info_message("info", f"Loading configuration file: {yaml_file_path}")
            with open(yaml_file_path, 'r') as file:
                depth = yaml.safe_load(file)

            module_path = Path(self.collector_dir) / f"{self.site_name}_{self.project_name}.py"

            class_name_in_site_script = f"{self.site_name}_{self.project_name}"
            class_name_in_site_script = ''.join([word.capitalize() for word in class_name_in_site_script.split('_')])
            try:
                spec = importlib.util.spec_from_file_location(class_name_in_site_script, module_path)
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                SiteClass = getattr(module, class_name_in_site_script)
                site_instance = SiteClass()
            except Exception as e:
                CommonModule.print_error_message("error", f"Error importing module from {module_path}: {e}")
                logging.exception("Module import failed")
                return

            seed_url = depth["depth0"]["seed_url"]

            self.get_final_url(seed_url, depth, 0, len(depth) - 1, site_instance)
            #self.enter_count_in_sheet()

        except Exception as e:
            CommonModule.print_error_message("error", f"Unhandled error during execution: {e}")
            logging.exception("Unhandled error during execution")
            raise

    def main(self):
        CommonModule.print_info_message("info", f"Starting script execution of url_collector for project: {self.project_name} and site: {self.site_name}")
        self.output_dir = Path(self.base_dir) / f"scrape_output/collector_output/{self.project_name}"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.collector_dir = Path(self.base_dir) / f"url_collector/{self.project_name}"
        filepath = self.output_dir / f"{self.site_name}_{self.project_name}.txt"
        with open(filepath, 'w') as file:
            file.write('')
        self.main_execution()

class UrlFetcher:
    def __init__(self, base_dir, project_name, site_name):
        CommonModule.print_info_message("info", f"Initializing UrlFetcher for project: {project_name} and site: {site_name}")
        self.base_dir = base_dir
        self.output_dir = ""
        self.project_name = project_name
        self.site_name = site_name

    def fetch_collector_output(self):
        CommonModule.print_info_message("info", f"Fetching collector output for project: {self.project_name} and site: {self.site_name}")
        urls = []
        try:
            self.output_dir = Path(self.base_dir) / f"scrape_output/collector_output/{self.project_name}"
            filepath = self.output_dir / f"{self.site_name}_{self.project_name}.txt"

            with open(filepath, "r") as f:
                for url in f:
                    urls.append(url.strip())
        except FileNotFoundError as e:
            CommonModule.print_error_message("error", f"File not found: {filepath}")
            logging.exception("File not found")
        return urls

    def main(self):
        CommonModule.print_info_message("info", f"Starting script execution of url_fetcher for project: {self.project_name} and site: {self.site_name}")
        yaml_file_path = Path(self.base_dir) / f"url_collector/{self.project_name}/{self.site_name}_{self.project_name}.yml"
        CommonModule.print_info_message("info", f"Loading configuration file: {yaml_file_path}")
        with open(yaml_file_path, 'r') as file:
            yaml_content = yaml.safe_load(file)

        extended_header = yaml_content.get("request_params", {}).get("extended_header", {})
        urls = self.fetch_collector_output()
        today = date.today()
        formatted_date = today.strftime('%Y%m%d')
        output_dir = Path(self.base_dir) / f"scrape_output/fetcher_output/{self.project_name}/{formatted_date}/{self.site_name}_{self.project_name}"
        output_dir.mkdir(parents=True, exist_ok=True)

        output_queue = Path(self.base_dir) / f"scrape_output/fetcher_output/{self.project_name}/{formatted_date}/{self.site_name}_{self.project_name}.txt"
        
        with open(output_queue, "a") as file:  # Open in append mode
            for key in urls:
                if not key:
                    continue
                url = key.split("|")[0]
                data = {}
                if extended_header:
                    result = CommonModule.get_page_content_hash(url, extended_header)
                else:
                    result = CommonModule.get_page_content_hash(url)
                data["url"] = key
                output_file = output_dir / f"{formatted_date}{CommonModule.encode(key)}.html"
                data["output_file"] = str(output_file)
                
                if result["status_code"] == 200:
                    with open(output_file, "wb") as f:
                        f.write(result["page_doc"].encode("utf-8"))
                    CommonModule.print_info_message("success", f"Successfully fetched page content for URL: {url}")
                else:
                    CommonModule.print_error_message("error", f"Failed to fetch page content for URL: {url}")
                
                # Write the data to the file immediately
                file.write(str(data) + "\n")

class UrlExtractor:
    def __init__(self, base_dir, project_name, site_name):
        CommonModule.print_info_message("info", f"Initializing UrlExtractor for project: {project_name} and site: {site_name}")
        self.base_dir = base_dir
        self.project_name = project_name
        self.site_name = site_name
        self.extractor_dir = Path(base_dir) / "url_data_extractor"
        self.count = 0

    def extract_records(self, output, page_doc, config, site_instance):
        CommonModule.print_info_message("info", f"Extracting records for project: {self.project_name} and site: {self.site_name}")
        """
        Modify and extract multiple records from the page_doc using site-specific methods.
        Args:
            page_doc: Parsed document for the fetched page.
            config: Field extraction rules from the YAML configuration.
            site_instance: Instance of the site-specific class.
        Returns:
            List of extracted records.
        """
        # Use the site-specific method to modify the page_doc
        subsections = site_instance.modify_page_doc(output, page_doc)

        records = []
        for sub_doc in subsections:
            record = {}
            for field, rules in config['fields'].items():
                method_name = f"get_{field}"
                if hasattr(site_instance, method_name):
                    extraction_method = getattr(site_instance, method_name)
                    try:
                        # Call the respective extraction method with the sub_doc and field rules
                        record[field] = extraction_method(sub_doc, {**rules, 'url': config.get('domain', '')})
                    except Exception as e:
                        record[field] = None
                        logging.warning(f"Error extracting field {field}: {e}")
                else:
                    logging.warning(f"Method {method_name} not implemented for field {field}.")
            records.append(record)
        CommonModule.print_info_message("success", f"Records extracted successfully for project: {self.project_name} and site: {self.site_name}")
        return records

    def main(self):
        CommonModule.print_info_message("info", f"Starting script execution of url_extractor for project: {self.project_name} and site: {self.site_name}")
        try:
            yaml_file_path = self.extractor_dir / f"{self.project_name}/{self.site_name}_{self.project_name}.yml"
            CommonModule.print_info_message("info", f"Loading configuration file: {yaml_file_path}")
            
            with open(yaml_file_path, 'r') as file:
                config = yaml.safe_load(file)  # Load the configuration from the YAML file

            module_path = self.extractor_dir / f"{self.project_name}/{self.site_name}_{self.project_name}.py"
            class_name_in_site_script = f"{self.site_name}_{self.project_name}"
            class_name_in_site_script = ''.join([word.capitalize() for word in class_name_in_site_script.split('_')])
            try:
                spec = importlib.util.spec_from_file_location(class_name_in_site_script, module_path)
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                SiteClass = getattr(module, class_name_in_site_script)
                site_instance = SiteClass()
            except Exception as e:
                CommonModule.print_error_message("error", f"Error importing module from {module_path}: {e}")
                return

            # Fetching file paths (where URLs are stored)
            today = date.today()
            formatted_date = today.strftime('%Y%m%d')
            output_queue = Path(self.base_dir) / f"scrape_output/fetcher_output/{self.project_name}/{formatted_date}/{self.site_name}_{self.project_name}.txt"
            with open(output_queue, 'r') as file:
                # Read all lines and strip any leading/trailing whitespaces
                file_paths = [line.strip() for line in file.readlines()]
            # Extract records for each file
            for output in file_paths:
                output_key = eval(output) 
                output = output_key.get("output_file")
                with open(output, 'r', encoding='utf-8') as file:
                    page_content = file.read()
                page_doc = etree.HTML(page_content)
                extracted_data = self.extract_records(output_key.get("url"),  page_doc, config, site_instance)
                # Write extracted data to CSV
                output_file = Path(self.base_dir) / f"scrape_output/extractor_output/{self.project_name}/{formatted_date}"
                output_file.mkdir(parents=True, exist_ok=True)
                output_file = output_file / f"{self.site_name}_{self.project_name}.csv"
                
                with open(output_file, mode='a', newline='', encoding='utf-8') as file:
                    writer = csv.DictWriter(file, fieldnames=config['fields'].keys())
                    if file.tell() == 0:
                        writer.writeheader()
                    writer.writerows(extracted_data)

        except Exception as e:
            CommonModule.print_error_message("error", f"Unhandled error during execution: {e}")
            logging.exception("Unhandled error during execution")
            raise

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python url_fetcher.py <project_name> <site_name>")
        sys.exit(1)
    method_to_execute = sys.argv[1]
    project_name = sys.argv[2]
    site_name = sys.argv[3]
    base_dir = "C:/Users/shanj/OneDrive/Desktop/web-scrapping-pipeline"
    if method_to_execute == "url_collector":
        url_collector = UrlCollector(base_dir,project_name,site_name)
        url_collector.main()
    elif method_to_execute == "url_fetcher":
        url_fetcher = UrlFetcher(base_dir,project_name,site_name)
        url_fetcher.main()
    elif method_to_execute == "url_extractor":
        url_extractor = UrlExtractor(base_dir,project_name,site_name)
        url_extractor.main()